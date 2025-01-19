import time
import openpyxl
import json

from pathlib import Path
from openpyxl.workbook import Workbook

from instruments import config
from instruments import data_instruments as DI
from instruments import data_scrappers as DS
from instruments import  import_generators as IG

# Initialisation of sheets
try:
    export_file = openpyxl.open("name.xlsx", False)  # Main file to work, Prom export table
    models_book = openpyxl.load_workbook("name.xlsx")  # Auto filled table (articule, engl, ru, ukr)
    data_changes = config.data_changes
    # region Sheet initialisaion
    with open("data/links_data.json", "r", encoding="utf-8") as file:
        links_data = json.load(file)

    export_sheet = export_file.active

    blank_book = openpyxl.open("data/Empty file for auto seat case.xlsx", False)  # Empty blank for any tasks, flexible
    blank_products = blank_book["Export Products Sheet"]
    blank_groups = blank_book["Export Groups Sheet"]

    book_empty = openpyxl.Workbook() # Empty table
    book_empty.remove(book_empty["Sheet"])
    empty_sheet = book_empty.create_sheet("Export Products Sheet")
    new_groups_sheet = book_empty.create_sheet("Export Groups Sheet")

    models_sheet = models_book["Export Products Sheet"]
    groups_sheet = models_book["Export Groups Sheet"]
    # endregion
except Exception as ex:
    print(ex)
    raise SystemExit


def main():
    start = time.time()
    # DI.init_project() # Initialises all dirs and files for work

    # Chose options
    # region data_scrappers
    # DS.large_import_data_to_excel("name.xlsx", "name", 1, export_sheet, empty_sheet, new_groups_sheet, book_empty)  # Useful for make table with names etc.(constructor)
    # DS.key_generator("name.xlsx", models_sheet, empty_sheet, book_empty) # Useful only for generating keys
    # endregion

    # region import_generators

    # endregion

    # region data_instruments
    # DI.clean_descriptions()         # Cleans all data from all descriptions files
    # DI.description_splitter()
    # endregion

    # region Closing sheets
    export_file.close()
    models_book.close()
    book_empty.close()
    blank_book.close()
    # endregion

    print(f"\nTime elapsed: {time.time() - start} seconds")

if __name__ == '__main__':
    main()