import openpyxl
import json
from instruments import config

class Resources:

    # Initialisation of sheets
    def __init__(self):
        try:
            self.export_file = openpyxl.open("name.xlsx", False)  # Main file to work, Prom export table
            self.models_book = openpyxl.load_workbook("name.xlsx")  # Auto filled table (articule, engl, ru, ukr)
            self.data_changes = config.data_changes
            # region Sheet initialisaion
            with open("data/links_data.json", "r", encoding="utf-8") as file:
                self.links_data = json.load(file)

            self.export_sheet = self.export_file.active

            self.blank_book = openpyxl.open("data/Empty file for auto seat case.xlsx",
                                       False)  # Empty blank for any tasks, flexible
            self.blank_products = self.blank_book["Export Products Sheet"]
            self.blank_groups = self.blank_book["Export Groups Sheet"]

            self.book_empty = openpyxl.Workbook()  # Empty table
            self.book_empty.remove(self.book_empty["Sheet"])
            self.empty_sheet = self.book_empty.create_sheet("Export Products Sheet")
            self.new_groups_sheet = self.book_empty.create_sheet("Export Groups Sheet")

            self.models_sheet = self.models_book["Export Products Sheet"]
            self.groups_sheet = self.models_book["Export Groups Sheet"]
            # endregion
        except json.decoder.JSONDecodeError as error:
            print("Error load json file. It is probably empty")
        except FileNotFoundError as error:
            print(error)
        except Exception as ex:
            print(ex)
            raise SystemExit


    # Close all opened xlsx books
    def close(self):
        # region Closing sheets
        self.export_file.close()
        self.models_book.close()
        self.book_empty.close()
        self.blank_book.close()
        # endregion