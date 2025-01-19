from pathlib2 import Path
from openpyxl.workbook import Workbook
import datetime
import re

from instruments import config

# Project initialisation
def init_project():
    def create_path(*files):
        for i in files:
            if not i.exists():
                if i.suffix:
                    i.touch(exist_ok=True)
                    print(f"File '{i}' created")
                else:
                    i.mkdir(exist_ok=True)
                    print(f"Directory '{i}' created")

    # data dir
    folder_path = Path("data")
    excel_path = folder_path / "Empty file for auto seat case.xlsx"
    json_path = folder_path / "links_data.json"

    create_path(folder_path, json_path)

    if not excel_path.exists():
        wb = Workbook()
        products_sheet = wb.active

        products_sheet.title = "Export Products Sheet"
        for id, name in config.PRODUCTS_COLUMNS.items():
            products_sheet.cell(1, id).value = name

        groups_sheet = wb.create_sheet("Export Groups Sheet")
        for id, name in config.GROUPS_COLUMNS.items():
            groups_sheet.cell(1, id).value = name

        wb.save(excel_path)
        print(f"File '{excel_path}' created")

    # Descriptions dir
    folder_path = Path("descriptions")
    description_file_ru = folder_path / "Description main ru.txt"
    description_file_ukr = folder_path / "Description main ukr.txt"
    folder_ru = folder_path / "ru"
    folder_ukr = folder_path / "ukr"
    models_ru = folder_ru / "models"
    models_ukr = folder_ukr / "models"

    create_path(folder_path, description_file_ru, description_file_ukr,
                folder_ru, folder_ukr, models_ru, models_ukr)

    for i in range(1, 31):
        create_path(folder_ru / f"{i}.txt", folder_ukr / f"{i}.txt")
    for i in range(1, 11):
        create_path(models_ru / f"{i}.txt", models_ukr / f"{i}.txt")

    # Lists of cars, Work result dir
    lists_of_cars = Path("lists of cars")
    work_result = Path("work result")
    create_path(lists_of_cars, work_result)

# region Useful instruments
def clean_descriptions():
    dir_path = Path("Descriptions")
    dir_ru = dir_path / "ru"
    dir_ukr = dir_path / "ukr"
    models_ru = dir_ru / "models"
    models_ukr = dir_ukr / "models"

    main_files = tuple(map(str, dir_path.glob("*.txt")))
    ru_files = tuple(map(str, dir_ru.glob("*.txt")))
    ukr_files = tuple(map(str, dir_ukr.glob("*.txt")))
    ru_models = tuple(map(str, models_ru.glob("*.txt")))
    ukr_models = tuple(map(str, models_ukr.glob("*.txt")))

    def cleaner(*lists):
        for files_list in lists:
            for i in files_list:
                with open(i, "w", encoding="utf-8"):
                    pass

    cleaner(main_files, ru_files, ukr_files, ru_models, ukr_models)

def description_splitter():
    with open("new descriptions.txt", "r", encoding="utf-8") as file:
        lines = file.read().split("\n")
        clean_lines = tuple(filter(lambda line: len(line) > 25, lines))

    for num, line in enumerate(clean_lines):
        with open(f"Descriptions/ru/{num + 1}.txt", "w", encoding="utf-8") as file:
            file.write(line)
        print(f"{num + 1}: {line}")

def how_many_marks(export_sheet):
    marks = get_all_marks(export_sheet)
    print(f"There are in total: {len(marks)} marks, not counting two-coloured positions (like skoda, skoda_white)")
    print(marks)

def check_duplicates(models_sheet):
    duplicates = []
    with open("dublicates_log.txt", "w", encoding="utf-8") as file:
        file.write(str(datetime.datetime.now()) + "\n\n")
        for row in range(1, models_sheet.max_row + 1):
            name = models_sheet.cell(row, 2).value
            name_add = models_sheet.cell(row, 5).value
            full_name = name + name_add
            if full_name in duplicates:
                info = f"{row}. {full_name}"
                print(info)
                file.write(info + "\n")
            else:
                duplicates.append(full_name)
# endregion

# region Data scrappers additions
# region large_import_data_to_excel
def create_group(mark, duplicates_groups):
    if mark not in duplicates_groups and mark is not None:
        duplicates_groups.append(mark)

    if mark is None:
        group_id = 1
    else:
        group_id = duplicates_groups.index(mark)

    return group_id, mark, duplicates_groups

def add_gift_keys(key_ru, key_ukr, name_ru, name_ukr):
    if len(key_ru) <= 1024 and key_ru.find("Подарок") == False:
        # Ru
        key_ru = key_ru + (f", Подарок владельцу автомобиля {name_ru}, Подарок водителю {name_ru}, "
                           f"Подарок в машину {name_ru}, Подарок для {name_ru}, Подарок для владельца {name_ru}")
        # Ukr
        key_ukr = key_ukr + (f", Подарунок власнику автомобіля {name_ukr}, Подарунок водієві {name_ukr}, "
                             f"Подарунок до авто {name_ukr}, Подарунок для {name_ukr}, Подарунок для власника {name_ukr}")
    return key_ru, key_ukr
# endregion

# region get_photo_data
def get_mark(row, sheet):
    for i in range(50, 85):
        if sheet.cell(row, i).value == "Марка":
            return sheet.cell(row, i + 2).value

def get_colour(row, sheet):
    for i in range(50, 85):
        if sheet.cell(row, i).value == "Цвет":
            return sheet.cell(row, i + 2).value

def get_all_marks(export_sheet):
    duplicates = []
    for row in range(2, export_sheet.max_row + 1):
        for i in range(50, 85):
            cell = export_sheet.cell(row, i).value
            if cell == "Марка":
                mark = export_sheet.cell(row, i + 2).value
                if mark not in duplicates:
                    duplicates.append(mark)
                break

    return duplicates

def create_empty_marks_coloured_dict(duplicates, colours):
    empty_dict = {}
    for i in duplicates:
        empty_dict.update([(i, {})])
        for colour in colours:
            sub_dict = empty_dict.get(i)
            sub_dict.update([(colour, "")])

    return empty_dict

def create_empty_coloured_dict(marks, colours):
    empty_dict = {}
    if len(colours) == 1:
        for mark in marks:
            empty_dict.update([(mark, "")])
    else:
        for colour in colours:
            empty_dict.update([(colour, "")])

    return empty_dict
# endregion
# endregion

# region Import generators additions
def get_groups_data(groups_sheet):
    groups = {}
    for row in range(1, groups_sheet.max_row + 1):
        group_id = groups_sheet.cell(row, 1).value
        name = groups_sheet.cell(row, 2).value
        groups.update([(name, group_id)])
    return groups

def fulfill_groups_data(blank_groups, groups_data, parent_group):
    group_names = tuple(groups_data.keys())
    group_ids = tuple(groups_data.values())
    print(group_names)
    print(group_ids)

    if parent_group:
        for row in range(3, len(group_ids) + 2):
            blank_groups.cell(row, 6).value = group_ids[0]

    for i in range(len(group_names)):
        blank_groups.cell(i + 2, 2).value = group_names[i]
        blank_groups.cell(i + 2, 4).value = group_ids[i]

def replace_names(full_name, name, name_engl, seat_year_name, colour, add_colour):
    if seat_year_name is None:
        full_name = full_name.replace(" name2", "")

    if add_colour:
        full_name = f"{full_name} {colour}"

    old_words = ("name", "name_en", "name2")
    new_words = ( name, name_engl, seat_year_name)
    for index, word in enumerate(old_words):
        pattern =  r"\b" + word + r"\b"
        full_name = re.sub(pattern, new_words[index], full_name)
    return full_name

def descriptions_generator(dcounter, new_name, language):
    with open(f"Descriptions/Description main {language}.txt", "r", encoding="utf-8") as file:
        data = file.read()
    with open(f"Descriptions/{language}/{dcounter}.txt", "r", encoding="utf-8") as file:
        data_text = file.read()
        new_text = data.replace("description", f"{data_text}")
        new_text = new_text.replace("name", f"{new_name}")
    return new_text
# endregion